import os
import time
import collections
from docx2python import docx2python  # pip install docx2python
import openpyxl  # pip install openpyxl
import fitz  # pip install PyMuPDF
# import win32com.client  # pip install pywin32 (this only works on Windows)

# .txt, .rtf, .xls
def binary_search(file_path, text):
    """Search for raw text encoded as UTF-8."""
    matches = []
    text = text.lower().encode('utf8')
    with open(file_path, 'rb') as file:
        data = file.read().lower()
    pos = data.find(text)
    while pos >= 0:
        matches.append(pos)
        pos = data.find(text, pos + 1)
    return matches


# .doc, .dot
def mbcs_search(file_path, text):
    """Search for raw text encoded as UTF-16."""
    matches = []
    utf16_text = text.lower().encode('utf-16')[2:]

    with open(file_path, 'rb') as file:
        data = file.read().lower()

    pos = data.find(utf16_text)
    while pos >= 0:
        matches.append(pos)
        pos = data.find(utf16_text, pos + len(utf16_text))

    return matches

def combined_search(file_path, text):
    """
    Perform both mbcs_search and binary_search on the file.
    """
    matches_mbcs = mbcs_search(file_path, text)
    matches_binary = binary_search(file_path, text)

    # Combine the results and remove duplicates
    combined_matches = list(set(matches_mbcs + matches_binary))

    return sorted(combined_matches)


# .docx, .docm
def extract_text_from_docx_element(element):
    """
    Recursively extracts text from a docx2python element,
    handling nested lists and concatenating text content.
    """
    if isinstance(element, list):
        return '\n'.join([extract_text_from_docx_element(item) for item in element])
    elif isinstance(element, str):
        return element
    else:
        return ''

def docx_python_search(file_path, text):
    try:
        result = docx2python(file_path)
        # Extracting text properly from the docx2python result object, handling nested structures
        text_content = extract_text_from_docx_element(result.body)
        text_content = text_content.lower()
        return [file_path] if text.lower() in text_content else []
    except Exception as e:
        print(f"Error processing file {file_path}: {e}")
        return []


# .xlsx
def xlsx_search(file_path, text):
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        matches = []
        for sheet in workbook.sheetnames:
            worksheet = workbook[sheet]
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value and text.lower() in str(cell.value).lower():
                        matches.append((sheet, cell.coordinate, cell.value))
        return matches
    except Exception as e:
        print(f"Error processing file {file_path}: {e}")
        return []


# .pdf
def pdf_search(file_path, text):
    try:
        doc = fitz.open(file_path)
        matches = []
        text_lower = text.lower()  # Convert search text to lowercase
        for i in range(len(doc)):
            page = doc.load_page(i)
            text_instances = page.search_for(text_lower, quads=True)  # Search for the lowercase text
            if text_instances:
                matches.append(i + 1)
        doc.close()
        return matches
    except Exception as e:
        print(f"Error processing file {file_path}: {e}")
        return []


if __name__ == "__main__":
    path = os.getcwd()
    text = input("Enter text to search for: ").lower()  # Convert input text to lowercase
    file_paths = []
    extensions = collections.Counter()

    for root, dirs, files in os.walk(path):
        for filename in files:
            file_paths.append(os.path.join(root, filename))
            basename, extension = os.path.splitext(filename)
            extensions[extension.lower()] += 1

    print(f"Found {len(file_paths)} files, {len(extensions)} unique extensions.")
    print("Unique extensions and counts:")
    for ext, count in extensions.items():
        print(f"\t{ext}: {count}")
    print("Searching...")
    start = time.time()

    for file_path in file_paths:
        matches = []
        extension = os.path.splitext(file_path)[1].lower()
        filename = os.path.basename(file_path)
        # Skip temporary files
        if filename.startswith('~$'):
            continue
        if extension.endswith(('.docx', '.docm')):
            matches = docx_python_search(file_path, text)
        elif extension.endswith('.xlsx'):
            matches = xlsx_search(file_path, text)
        elif extension.endswith('.txt'):
            matches = binary_search(file_path, text)
        elif extension.endswith('.rtf'):
            matches = binary_search(file_path, text)
        elif extension.endswith('.xls'):
            matches = binary_search(file_path, text)
        elif extension.endswith(('.dot', '.doc')):
            # matches = mbcs_search(file_path, text)
            matches = combined_search(file_path, text)
        elif extension == '.pdf':
            matches = pdf_search(file_path, text)

        elif not extension.endswith(('', '.jpg', '.db', '.png', '.wbk', '.py',
                                '.vsd', '.jpeg', '.pptx', '.shs',
                                '.lnk', '.tmp', '.bmp')):
            print("\t'{}' skipped".format(file_path))

        if matches:
            match_details = f"{len(matches)} matches in '{file_path}'"
            if extension == '.pdf':
                match_details += f", Pages: {', '.join(map(str, matches))}"
            print(match_details)

    elapsed = time.time() - start
    print(f"Finished in {elapsed:.3f} seconds.")